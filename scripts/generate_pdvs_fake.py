"""
Generates a fake but realistic PDV (point of sale) spreadsheet for Bang Bang.
Used as a simulation dataset until the real 287-PDV list is provided.

Distribution: 287 PDVs across 12 Brazilian states, leaving 15 states gray
to demonstrate the smart-map feature (only active states get colored).
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

random.seed(42)  # deterministic output

# ---------- City data: (city, UF, cep_prefix, lat, lng, weight) ----------
CITIES = {
    "SP": [
        ("São Paulo",      "SP", "01", -23.5505, -46.6333, 60),
        ("Santos",         "SP", "11", -23.9608, -46.3336, 5),
        ("Campinas",       "SP", "13", -22.9099, -47.0626, 5),
        ("Guarulhos",      "SP", "07", -23.4628, -46.5333, 3),
        ("Santo André",    "SP", "09", -23.6633, -46.5383, 3),
        ("São Bernardo do Campo", "SP", "09", -23.6914, -46.5646, 3),
        ("Osasco",         "SP", "06", -23.5320, -46.7917, 2),
        ("Ribeirão Preto", "SP", "14", -21.1775, -47.8103, 2),
        ("Sorocaba",       "SP", "18", -23.5015, -47.4526, 1),
        ("Jundiaí",        "SP", "13", -23.1857, -46.8978, 1),
    ],
    "RJ": [
        ("Rio de Janeiro", "RJ", "20", -22.9068, -43.1729, 35),
        ("Niterói",        "RJ", "24", -22.8833, -43.1036, 5),
        ("Petrópolis",     "RJ", "25", -22.5050, -43.1786, 3),
        ("Nova Iguaçu",    "RJ", "26", -22.7592, -43.4511, 3),
        ("Duque de Caxias","RJ", "25", -22.7858, -43.3119, 2),
        ("São Gonçalo",    "RJ", "24", -22.8267, -43.0538, 2),
    ],
    "MG": [
        ("Belo Horizonte", "MG", "30", -19.9167, -43.9345, 20),
        ("Contagem",       "MG", "32", -19.9317, -44.0536, 4),
        ("Uberlândia",     "MG", "38", -18.9113, -48.2622, 3),
        ("Juiz de Fora",   "MG", "36", -21.7642, -43.3503, 3),
        ("Betim",          "MG", "32", -19.9680, -44.1984, 2),
    ],
    "PR": [
        ("Curitiba",       "PR", "80", -25.4284, -49.2733, 15),
        ("Londrina",       "PR", "86", -23.3045, -51.1696, 4),
        ("Maringá",        "PR", "87", -23.4253, -51.9386, 2),
        ("Ponta Grossa",   "PR", "84", -25.0950, -50.1619, 2),
        ("Foz do Iguaçu",  "PR", "85", -25.5478, -54.5882, 1),
    ],
    "RS": [
        ("Porto Alegre",   "RS", "90", -30.0346, -51.2177, 12),
        ("Caxias do Sul",  "RS", "95", -29.1678, -51.1794, 4),
        ("Pelotas",        "RS", "96", -31.7719, -52.3425, 2),
        ("Santa Maria",    "RS", "97", -29.6842, -53.8069, 2),
    ],
    "SC": [
        ("Florianópolis",  "SC", "88", -27.5954, -48.5480, 8),
        ("Joinville",      "SC", "89", -26.3044, -48.8487, 4),
        ("Blumenau",       "SC", "89", -26.9194, -49.0661, 3),
        ("Chapecó",        "SC", "89", -27.1000, -52.6152, 2),
        ("Criciúma",       "SC", "88", -28.6775, -49.3697, 1),
    ],
    "BA": [
        ("Salvador",              "BA", "40", -12.9777, -38.5016, 9),
        ("Feira de Santana",      "BA", "44", -12.2664, -38.9663, 3),
        ("Vitória da Conquista",  "BA", "45", -14.8615, -40.8442, 2),
    ],
    "DF": [
        ("Brasília",   "DF", "70", -15.7801, -47.9292, 7),
        ("Taguatinga", "DF", "72", -15.8333, -48.0553, 3),
        ("Ceilândia",  "DF", "72", -15.8194, -48.1073, 2),
    ],
    "GO": [
        ("Goiânia",  "GO", "74", -16.6869, -49.2648, 7),
        ("Anápolis", "GO", "75", -16.3281, -48.9531, 3),
    ],
    "PE": [
        ("Recife",   "PE", "50", -8.0476, -34.8770, 5),
        ("Olinda",   "PE", "53", -8.0089, -34.8553, 2),
        ("Jaboatão dos Guararapes", "PE", "54", -8.1127, -35.0148, 1),
        ("Caruaru",  "PE", "55", -8.2845, -35.9760, 1),
    ],
    "ES": [
        ("Vitória",    "ES", "29", -20.3155, -40.3128, 4),
        ("Vila Velha", "ES", "29", -20.3297, -40.2925, 2),
        ("Serra",      "ES", "29", -20.1295, -40.3076, 1),
    ],
    "CE": [
        ("Fortaleza", "CE", "60", -3.7319, -38.5267, 4),
        ("Caucaia",   "CE", "61", -3.7360, -38.6533, 2),
    ],
}

# Quota per state (totals 287)
QUOTAS = {
    "SP": 85, "RJ": 50, "MG": 32, "PR": 24, "RS": 20, "SC": 18,
    "BA": 14, "DF": 12, "GO": 10, "PE": 9, "ES": 7, "CE": 6,
}
assert sum(QUOTAS.values()) == 287

# ---------- Establishment data ----------
TIPOS = ["Bar", "Adega", "Mercado", "Conveniência", "Restaurante",
         "Hamburgueria", "Empório", "Pub", "Distribuidora", "Lanchonete"]

TIPO_WEIGHTS = [25, 20, 15, 10, 10, 8, 5, 3, 2, 2]  # sums 100

PREFIXOS_NOME = [
    "Bar do", "Boteco do", "Adega", "Empório", "Mercado",
    "Armazém", "Distribuidora", "Casa do", "Ponto do", "Espaço",
]

NOMES_PROPRIOS = [
    "Zé", "João", "Pedro", "Tião", "Vitinho", "Marquinhos", "Ricardão",
    "Seu Chico", "Toninho", "Paulinho", "Juca", "Bira", "Binho",
    "Caco", "Dudu", "Gustavo", "Nando", "Rogério", "Mário",
]

NOMES_FANTASIA = [
    "Barril", "Boteco 157", "Mistura Fina", "Pé Sujo", "Esquina Feliz",
    "Bebidas Premium", "Cervejaria Central", "Adega do Vale", "Empório Verde",
    "Mercadinho Central", "Conveniência 24h", "Estação Gelada", "Bar do Rock",
    "Copo Meio Cheio", "Hora Feliz", "Pitstop", "Balcão 77", "Gelo Seco",
    "Rua 15", "Travessa Bar", "Oitavo Andar", "Quintal", "Térreo",
    "Bar do Mercado", "Vila Madalena Bar", "Estação Bebidas",
]

TIPOS_LOGRADOURO = ["Rua", "Avenida", "Alameda", "Travessa", "Praça", "Estrada"]

NOMES_LOGRADOURO = [
    "das Flores", "Brasil", "São João", "Sete de Setembro", "Independência",
    "da Consolação", "Augusta", "Paulista", "Atlântica", "Rio Branco",
    "das Américas", "Getúlio Vargas", "Presidente Vargas", "XV de Novembro",
    "João Pessoa", "dos Andradas", "Conselheiro", "Senador Queiroz",
    "Bela Cintra", "Oscar Freire", "Haddock Lobo", "Nestor Pestana",
    "Treze de Maio", "Voluntários da Pátria", "Cardeal Arcoverde",
    "Aspicuelta", "Wisard", "Harmonia", "Fradique Coutinho",
]

BAIRROS = {
    "SP": ["Vila Madalena", "Pinheiros", "Jardins", "Itaim Bibi", "Moema",
           "Vila Olímpia", "Perdizes", "Higienópolis", "Bela Vista", "Liberdade",
           "Consolação", "República", "Santa Cecília", "Barra Funda", "Lapa",
           "Tatuapé", "Mooca", "Ipiranga", "Vila Mariana", "Saúde"],
    "RJ": ["Copacabana", "Ipanema", "Leblon", "Botafogo", "Flamengo",
           "Laranjeiras", "Tijuca", "Barra da Tijuca", "Jardim Botânico",
           "Lapa", "Santa Teresa", "Urca", "Gávea", "Humaitá"],
    "MG": ["Savassi", "Funcionários", "Lourdes", "Centro", "Santa Tereza",
           "Pampulha", "Floresta", "Anchieta", "São Pedro", "Belvedere"],
    "PR": ["Batel", "Centro", "Água Verde", "Bigorrilho", "Cabral",
           "Juvevê", "Rebouças", "Portão", "Santa Felicidade", "Mercês"],
    "RS": ["Moinhos de Vento", "Centro Histórico", "Bom Fim", "Cidade Baixa",
           "Menino Deus", "Petrópolis", "Bela Vista", "Rio Branco", "Higienópolis"],
    "SC": ["Centro", "Trindade", "Lagoa da Conceição", "Campeche",
           "Ingleses", "Canasvieiras", "Jurerê", "Córrego Grande"],
    "BA": ["Pelourinho", "Barra", "Ondina", "Rio Vermelho", "Pituba",
           "Graça", "Itapuã", "Stella Maris", "Itaigara"],
    "DF": ["Asa Sul", "Asa Norte", "Lago Sul", "Lago Norte", "Sudoeste",
           "Cruzeiro", "Noroeste", "Setor Bancário Sul"],
    "GO": ["Setor Bueno", "Setor Marista", "Setor Oeste", "Centro",
           "Jardim Goiás", "Setor Sul", "Alto da Glória"],
    "PE": ["Boa Viagem", "Pina", "Casa Forte", "Espinheiro", "Graças",
           "Derby", "Recife Antigo", "Poço da Panela"],
    "ES": ["Praia do Canto", "Jardim da Penha", "Enseada do Suá", "Centro",
           "Bento Ferreira", "Santa Lúcia", "Mata da Praia"],
    "CE": ["Meireles", "Aldeota", "Mucuripe", "Iracema", "Varjota",
           "Papicu", "Cocó", "Dionísio Torres"],
}

REPRESENTANTES = [
    "Ricardo Almeida",  "Juliana Costa",   "Marcelo Rocha",    "Patrícia Souza",
    "Fernando Lima",    "Amanda Ferreira", "Gabriel Martins",  "Beatriz Nogueira",
    "Thiago Moreira",   "Carolina Duarte", "Henrique Barbosa", "Larissa Pinto",
]

HORARIOS = [
    "Seg-Sáb 17h-02h",
    "Ter-Dom 18h-01h",
    "Seg-Dom 11h-23h",
    "Qua-Dom 19h-03h",
    "Seg-Sáb 10h-22h",
    "Seg-Dom 08h-00h",
    "Seg-Sex 11h-23h | Sáb-Dom 11h-01h",
    "Ter-Sáb 18h-02h | Dom 14h-22h",
]

PLATAFORMAS_DELIVERY = [
    ("iFood",  "https://www.ifood.com.br/delivery/"),
    ("Rappi",  "https://www.rappi.com.br/stores/"),
    ("Zé Delivery", "https://www.ze.delivery/"),
]


def generate_name(tipo: str) -> str:
    """Generate a plausible establishment name."""
    r = random.random()
    if tipo in ("Bar", "Boteco", "Pub", "Hamburgueria", "Lanchonete"):
        if r < 0.35:
            return f"{random.choice(['Bar do','Boteco do','Bar','Boteco'])} {random.choice(NOMES_PROPRIOS)}"
        if r < 0.75:
            return random.choice(NOMES_FANTASIA)
        return f"{tipo} {random.choice(NOMES_FANTASIA)}"
    if tipo in ("Adega", "Empório", "Mercado", "Distribuidora", "Conveniência"):
        if r < 0.4:
            return f"{tipo} {random.choice(NOMES_PROPRIOS)}"
        if r < 0.75:
            return f"{tipo} {random.choice(NOMES_FANTASIA)}"
        return random.choice(NOMES_FANTASIA)
    return random.choice(NOMES_FANTASIA)


def generate_cep(prefix: str) -> str:
    suffix_a = random.randint(0, 999)
    suffix_b = random.randint(0, 999)
    return f"{prefix}{suffix_a:03d}-{suffix_b:03d}"


def generate_address() -> tuple[str, int]:
    tipo = random.choice(TIPOS_LOGRADOURO)
    nome = random.choice(NOMES_LOGRADOURO)
    numero = random.randint(20, 3500)
    return f"{tipo} {nome}", numero


def generate_phone(uf: str) -> str:
    ddd_map = {
        "SP": ["11", "12", "13", "14", "15", "16", "17", "18", "19"],
        "RJ": ["21", "22", "24"], "MG": ["31", "32", "33", "34", "35", "37", "38"],
        "PR": ["41", "42", "43", "44", "45", "46"],
        "RS": ["51", "53", "54", "55"], "SC": ["47", "48", "49"],
        "BA": ["71", "73", "74", "75", "77"], "DF": ["61"],
        "GO": ["62", "64"], "PE": ["81", "87"], "ES": ["27", "28"],
        "CE": ["85", "88"],
    }
    ddd = random.choice(ddd_map[uf])
    num = f"9{random.randint(1000,9999)}-{random.randint(1000,9999)}"
    return f"({ddd}) {num}"


def jitter(v: float, amount: float = 0.045) -> float:
    return round(v + random.uniform(-amount, amount), 6)


def build_rows() -> list[dict]:
    rows: list[dict] = []
    pdv_id = 1
    start_date = datetime(2024, 3, 1)

    for uf, quota in QUOTAS.items():
        city_pool = CITIES[uf]
        total_weight = sum(c[5] for c in city_pool)

        for _ in range(quota):
            r = random.uniform(0, total_weight)
            acc = 0
            chosen = city_pool[0]
            for city in city_pool:
                acc += city[5]
                if r <= acc:
                    chosen = city
                    break
            city_name, uf_code, cep_prefix, lat, lng, _ = chosen

            tipo = random.choices(TIPOS, weights=TIPO_WEIGHTS, k=1)[0]
            nome = generate_name(tipo)
            logradouro, numero = generate_address()
            bairro = random.choice(BAIRROS[uf])
            cep = generate_cep(cep_prefix)
            plat_lat = jitter(lat)
            plat_lng = jitter(lng)
            telefone = generate_phone(uf)
            horario = random.choice(HORARIOS)
            gmaps = f"https://www.google.com/maps/search/?api=1&query={plat_lat},{plat_lng}"
            delivery = ""
            if random.random() < 0.55:
                plat_name, plat_url = random.choice(PLATAFORMAS_DELIVERY)
                slug = nome.lower().replace(" ", "-").replace("ç", "c").replace("ã", "a")
                delivery = f"{plat_url}{slug}"
            ativo = "SIM" if random.random() > 0.05 else "NAO"
            rep = random.choice(REPRESENTANTES)
            data_cad = start_date + timedelta(days=random.randint(0, 760))
            obs = ""
            if random.random() < 0.15:
                obs = random.choice([
                    "PDV âncora da região",
                    "Alta rotatividade de estoque",
                    "Público jovem 18-30",
                    "Participou do lançamento",
                    "Foco em eventos e música ao vivo",
                    "Pede reposição semanal",
                    "Visibilidade premium na vitrine",
                ])

            rows.append({
                "id": f"PDV-{pdv_id:04d}",
                "nome": nome,
                "tipo": tipo,
                "endereco": f"{logradouro}, {numero}",
                "bairro": bairro,
                "cidade": city_name,
                "uf": uf_code,
                "cep": cep,
                "latitude": plat_lat,
                "longitude": plat_lng,
                "telefone": telefone,
                "horario_funcionamento": horario,
                "link_google_maps": gmaps,
                "link_delivery": delivery,
                "ativo": ativo,
                "representante": rep,
                "observacoes": obs,
                "data_cadastro": data_cad.strftime("%Y-%m-%d"),
            })
            pdv_id += 1

    random.shuffle(rows)
    # Reassign IDs sequentially after shuffle so they read in order
    for i, row in enumerate(rows, start=1):
        row["id"] = f"PDV-{i:04d}"
    return rows


def write_workbook(rows: list[dict], out_path: str) -> None:
    wb = Workbook()

    # ---------- Sheet 1: PDVs ----------
    ws = wb.active
    ws.title = "PDVs"

    headers = [
        ("id", "ID", 11),
        ("nome", "Nome do Estabelecimento", 32),
        ("tipo", "Tipo", 14),
        ("endereco", "Endereço", 34),
        ("bairro", "Bairro", 22),
        ("cidade", "Cidade", 22),
        ("uf", "UF", 6),
        ("cep", "CEP", 12),
        ("latitude", "Latitude", 12),
        ("longitude", "Longitude", 12),
        ("telefone", "Telefone", 16),
        ("horario_funcionamento", "Horário de Funcionamento", 32),
        ("link_google_maps", "Link Google Maps", 32),
        ("link_delivery", "Link Delivery", 32),
        ("ativo", "Ativo", 8),
        ("representante", "Representante Responsável", 26),
        ("observacoes", "Observações", 36),
        ("data_cadastro", "Data de Cadastro", 16),
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A1A1A")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_font = Font(name="Arial", size=10)
    body_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="E5E5E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_key, label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _label, _w) in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border
            if key in ("latitude", "longitude"):
                cell.number_format = "0.000000"

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # ---------- Sheet 2: Resumo (live formulas) ----------
    ws2 = wb.create_sheet("Resumo por Estado")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    resumo_headers = ["UF", "Estado", "Total PDVs", "Ativos", "% do Total"]
    for col_idx, label in enumerate(resumo_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws2.row_dimensions[1].height = 28

    uf_names = {
        "AC": "Acre", "AL": "Alagoas", "AM": "Amazonas", "AP": "Amapá",
        "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal", "ES": "Espírito Santo",
        "GO": "Goiás", "MA": "Maranhão", "MG": "Minas Gerais", "MS": "Mato Grosso do Sul",
        "MT": "Mato Grosso", "PA": "Pará", "PB": "Paraíba", "PE": "Pernambuco",
        "PI": "Piauí", "PR": "Paraná", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
        "RO": "Rondônia", "RR": "Roraima", "RS": "Rio Grande do Sul", "SC": "Santa Catarina",
        "SE": "Sergipe", "SP": "São Paulo", "TO": "Tocantins",
    }
    all_ufs_sorted = sorted(uf_names.keys())
    total_rows = len(rows)
    pdv_range_uf = f"PDVs!G2:G{total_rows + 1}"
    pdv_range_ativo = f"PDVs!O2:O{total_rows + 1}"

    for i, uf in enumerate(all_ufs_sorted, start=2):
        ws2.cell(row=i, column=1, value=uf).font = body_font
        ws2.cell(row=i, column=2, value=uf_names[uf]).font = body_font
        ws2.cell(row=i, column=3, value=f'=COUNTIF({pdv_range_uf},"{uf}")').font = body_font
        ws2.cell(row=i, column=4,
                 value=f'=COUNTIFS({pdv_range_uf},"{uf}",{pdv_range_ativo},"SIM")').font = body_font
        ws2.cell(row=i, column=5, value=f"=IFERROR(C{i}/$C$30,0)").font = body_font
        ws2.cell(row=i, column=5).number_format = "0.0%"
        for c in range(1, 6):
            ws2.cell(row=i, column=c).border = border
            ws2.cell(row=i, column=c).alignment = Alignment(horizontal="left", vertical="center")

    total_row = len(all_ufs_sorted) + 2
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=total_row, column=2, value="Brasil").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})").font = Font(name="Arial", bold=True)
    ws2.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})").font = Font(name="Arial", bold=True)
    ws2.cell(row=total_row, column=5, value=f"=IFERROR(C{total_row}/C{total_row},0)").font = Font(name="Arial", bold=True)
    ws2.cell(row=total_row, column=5).number_format = "0.0%"
    for c in range(1, 6):
        ws2.cell(row=total_row, column=c).fill = PatternFill("solid", start_color="F4F4F4")
        ws2.cell(row=total_row, column=c).border = border

    ws2.freeze_panes = "A2"

    # ---------- Sheet 3: Dicionário de Dados ----------
    ws3 = wb.create_sheet("Dicionário de Dados")
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 70

    dict_headers = ["Campo", "Tipo", "Obrigatório", "Descrição e Regras"]
    for col_idx, label in enumerate(dict_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
    ws3.row_dimensions[1].height = 28

    dictionary = [
        ("id", "Texto", "Sim", "Identificador único. Formato PDV-0001, sequencial. Gerado pelo sistema."),
        ("nome", "Texto", "Sim", "Nome comercial do estabelecimento como aparece na fachada."),
        ("tipo", "Enum", "Sim", "Categoria do PDV. Valores aceitos: Bar, Adega, Mercado, Conveniência, Restaurante, Hamburgueria, Empório, Pub, Distribuidora, Lanchonete."),
        ("endereco", "Texto", "Sim", "Logradouro e número. Ex.: Rua Augusta, 1200."),
        ("bairro", "Texto", "Sim", "Bairro do estabelecimento."),
        ("cidade", "Texto", "Sim", "Município. Grafia oficial (com acentos)."),
        ("uf", "Texto(2)", "Sim", "Sigla do estado em maiúsculas. Ex.: SP, RJ, MG."),
        ("cep", "Texto", "Sim", "CEP no formato 00000-000."),
        ("latitude", "Decimal", "Sim", "Coordenada geográfica. 6 casas decimais. Valores negativos no Brasil."),
        ("longitude", "Decimal", "Sim", "Coordenada geográfica. 6 casas decimais. Valores negativos no Brasil."),
        ("telefone", "Texto", "Não", "Formato (DDD) XXXXX-XXXX. Preferencialmente WhatsApp."),
        ("horario_funcionamento", "Texto", "Não", "Formato livre. Ex.: Seg-Sáb 17h-02h."),
        ("link_google_maps", "URL", "Não", "Gerado automaticamente a partir de lat/lng se não informado."),
        ("link_delivery", "URL", "Não", "Link direto para iFood, Rappi, Zé Delivery ou próprio."),
        ("ativo", "SIM/NAO", "Sim", "Controla se o PDV aparece no site. NAO = oculto, não exclui histórico."),
        ("representante", "Texto", "Não", "Nome do representante comercial responsável. Uso interno."),
        ("observacoes", "Texto", "Não", "Notas internas sobre o PDV. Não aparece no site."),
        ("data_cadastro", "Data", "Sim", "Formato YYYY-MM-DD. Data em que o PDV entrou na base."),
    ]
    for i, (field, tp, req, desc) in enumerate(dictionary, start=2):
        ws3.cell(row=i, column=1, value=field).font = Font(name="Arial", size=10, bold=True)
        ws3.cell(row=i, column=2, value=tp).font = body_font
        ws3.cell(row=i, column=3, value=req).font = body_font
        ws3.cell(row=i, column=4, value=desc).font = body_font
        for c in range(1, 5):
            ws3.cell(row=i, column=c).border = border
            ws3.cell(row=i, column=c).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws3.row_dimensions[i].height = 32

    ws3.freeze_panes = "A2"

    # ---------- Sheet 4: Instruções ----------
    ws4 = wb.create_sheet("Instruções")
    ws4.column_dimensions["A"].width = 100

    instrucoes = [
        ("Planilha de PDVs Bang Bang — Dataset Simulado", "title"),
        ("", ""),
        ("Este arquivo é um dataset FAKE gerado para desenvolvimento da dobra \"Onde Comprar\".", "body"),
        ("287 PDVs distribuídos em 12 estados. Os 15 estados restantes ficam \"cinzas\" no mapa.", "body"),
        ("", ""),
        ("Como substituir pelos dados reais", "h2"),
        ("1. Mantenha a mesma estrutura de colunas da aba \"PDVs\" — o site lê essa estrutura.", "body"),
        ("2. Mantenha os nomes dos campos no cabeçalho exatamente como estão (maiúsculas, acentos).", "body"),
        ("3. Preencha lat/lng com 6 casas decimais. Se não tiver, deixe vazio e avise — rodamos geocoding.", "body"),
        ("4. UF sempre em 2 letras maiúsculas. CEP sempre com hífen (00000-000).", "body"),
        ("5. A coluna \"ativo\" controla a visibilidade no site: SIM aparece, NAO oculta.", "body"),
        ("", ""),
        ("Como o mapa inteligente funciona", "h2"),
        ("O site lê a coluna UF dessa planilha e gera automaticamente a lista de estados ativos.", "body"),
        ("Estados com pelo menos 1 PDV ativo ficam coloridos no mapa. Os demais ficam cinza.", "body"),
        ("Adicionou uma loja em TO? No próximo build, TO acende. Zero manutenção no código.", "body"),
        ("", ""),
        ("Distribuição atual (dataset fake)", "h2"),
        ("SP: 85 | RJ: 50 | MG: 32 | PR: 24 | RS: 20 | SC: 18", "body"),
        ("BA: 14 | DF: 12 | GO: 10 | PE: 9 | ES: 7 | CE: 6", "body"),
        ("Estados cinzas: AC, AL, AM, AP, MA, MS, MT, PA, PB, PI, RN, RO, RR, SE, TO", "body"),
        ("", ""),
        ("Campos obrigatórios mínimos", "h2"),
        ("id, nome, tipo, endereco, bairro, cidade, uf, cep, latitude, longitude, ativo, data_cadastro", "body"),
        ("Ver aba \"Dicionário de Dados\" para detalhes de cada campo.", "body"),
    ]
    for i, (text, kind) in enumerate(instrucoes, start=1):
        cell = ws4.cell(row=i, column=1, value=text)
        if kind == "title":
            cell.font = Font(name="Arial", bold=True, size=16, color="1A1A1A")
            ws4.row_dimensions[i].height = 28
        elif kind == "h2":
            cell.font = Font(name="Arial", bold=True, size=12, color="1A1A1A")
            ws4.row_dimensions[i].height = 22
        else:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Reorder: put Instruções first for visibility
    wb.move_sheet("Instruções", offset=-3)

    wb.save(out_path)


if __name__ == "__main__":
    rows = build_rows()
    out = "/sessions/busy-adoring-ramanujan/mnt/bangbang-site/data/pdvs_bang_bang.xlsx"
    write_workbook(rows, out)
    print(f"OK: {len(rows)} PDVs written to {out}")
